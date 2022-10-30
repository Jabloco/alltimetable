import re
from datetime import datetime

import openpyxl

from pydantic import ValidationError


from files_names import contacts_file
from files_names import monitoring_file
from files_names import alltime_file

from pydentic_models import PhoneNum
from pydentic_models import MainData
from pydentic_models import FiscalData
from pydentic_models import DevicesData
from pydentic_models import EgaisData
from pydentic_models import ShopData
from pydentic_models import AllShopsInfo


def separate_num(words_list: list) -> tuple[list, list]:
    """
    Функция для разделения корпоративных и личных номеров телефона.

    Принимае список.

    Возвращает два списка person_phone и corp_phone. Именно в этом порядке
    """
    person_phone = []
    corp_phone = []
    # проходим по списку, ищем номера, отделяем корпоративный от личных
    for word in words_list:
        if not re.match('[\d]+', word):
            continue
        if (re.search('[личн]', word)
                and len(re.match('[\d]+', word).group(0)) == 11):
            person_phone.append(int(re.match('[\d]+', word).group(0)))
        elif (len(re.match('[\d]+', word).group(0)) == 4
                or len(re.match('[\d]+', word).group(0)) == 11):
            corp_phone.append(int(word))
    return person_phone, corp_phone


def find_yes(cell_data) -> bool:
    """
    Функция для нахождения ДА в ячейке.

    Возвращает True или False.
    """
    try:
        if re.match('(да)', cell_data, re.IGNORECASE).group(0):
            is_yes = True
    except AttributeError:
        is_yes = False
    except TypeError:
        is_yes = False
    return is_yes


def find_status(cell_data) -> bool:
    try:
        if re.match('(действующий)', cell_data, re.IGNORECASE).group(0):
            status = True
    except AttributeError:
        status = False
    except TypeError:
        status = False
    return status


def is_date(cell_data) -> datetime | None:
    """
    Функция для определения что данные в ячейке есть дата

    Возвращает datetime
    """
    if isinstance(cell_data, datetime):
        return cell_data
    else:
        return None


def find_shop_num(cell_data) -> int | None:
    try:
        if re.match('(маг[ ]+№|магазин[ ]+№)[\d]+', cell_data, re.IGNORECASE).group(0):
            return int(re.search('[\d]+', cell_data).group())
    except TypeError:
        return None
    except AttributeError:
        return None


def is_kkt_num(cell_data) -> str | None:
    """
    Функция для определения являются ли данные в ячейке каким либо номером кассы

    Заводской номер и рег.номер ККТ имеют одинаковую длину.

    Что это будет за номер зависит только от столбца
    """
    try:
        return re.match('[\d]{16}', cell_data).group(0)
    except TypeError:
        return None
    except AttributeError:
        return None


def find_post_index(cell_data) -> str | None:
    try:
        return re.match('[\d]{6}', str(cell_data)).group(0)
    except TypeError:
        return None
    except AttributeError:
        return None


def find_shop_kpp(cell_data) -> str | None:
    try:
        return re.search('[\d]{9}', str(cell_data)).group(0)
    except TypeError:
        return None
    except AttributeError:
        return None


def find_fn_period(cell_data) -> int | None:
    try:
        return int(re.search('13|15|36', cell_data).group(0))
    except TypeError:
        return None
    except AttributeError:
        return None


def find_fsrar_id(cell_data) -> str | None:
    try:
        return re.match('[\d]{12}', cell_data).group(0)
    except TypeError:
        return None
    except AttributeError:
        return None


def find_logic_num(cell_data) -> int | None:
    try:
        return re.match('[\d]{1,3}', cell_data).group(0)
    except TypeError:
        return None
    except AttributeError:
        return None


def find_shtrih_ver(cell_data) -> str | None:
    try:
        return re.search('(\d.\d.\d.\d)', cell_data).group(0)
    except AttributeError:
        return None
    except TypeError:
        return None


def contacts_parser(file_link: str) -> AllShopsInfo:
    """
    Из файла парсим номер магазина, номер партии, телефоны
    """
    contacts_book = openpyxl.load_workbook(file_link)
    worksheet = contacts_book.active
    all_shops = []
    for row in range(1, worksheet.max_row):
        for col in worksheet.iter_cols(2, worksheet.max_column):
            match col[row].column:
                case 2:  # номер маршрута
                    num_shipment_raw = re.findall('[0-9]+', str(col[row].value))
                case 3:  # номер магазина
                    shop_num_raw = col[row].value
                case 6:  # номера телефонов
                    if col[row].value:  # если ячейка не пустая
                        # разбираем ячейку на "слова", формируем список
                        words_list = re.findall('[\w]+', col[row].value)
                    person_phone, corp_phone = separate_num(words_list)
                    try:
                        phones = PhoneNum(
                            person_num=person_phone,
                            corp_num=corp_phone
                            )
                    except ValidationError as e:
                        print(e.json())

        try:
            shop_main_info = MainData(
                shop_shipment_num=num_shipment_raw,
                shop_num=shop_num_raw,
                shop_phone_num=phones
            )
        except ValidationError as e:
            print(e.json())

        try:
            shop_info = ShopData(
                main_info=shop_main_info
            )
        except ValidationError as e:
            print(e.json())
        all_shops.append(shop_info)
    return AllShopsInfo(shops=all_shops)


def alltimetable_parser(file_link) -> AllShopsInfo:
    alltimetable_book = openpyxl.load_workbook(file_link)
    worksheet = alltimetable_book.active
    all_shops = []
    for row in range(1, worksheet.max_row):
        for col in worksheet.iter_cols(2, worksheet.max_column):
            match col[row].column:
                case 2:  # номер магазина
                    shop_num_raw = find_shop_num(col[row].value)
                case 3:  # адрес магазина
                    shop_address_raw = col[row].value
                case 4:  # статус магазина
                    shop_status_raw = find_status(col[row].value)
                case 5:  # юридическое лицо
                    shop_entity_raw = col[row].value
                case 6:  # почтовый индекс
                    shop_post_index_raw = find_post_index(col[row].value)
                case 7:  # КПП
                    shop_kpp_raw = find_shop_kpp(col[row].value)
                case 9:  # модель фискальника
                    fiscal_model_raw = col[row].value
                case 10:  # имя в Такскоме
                    fiscal_taxcom_name_raw = col[row].value
                case 11:  # заводской номер ккт
                    fiscal_fabric_raw = is_kkt_num(col[row].value)
                case 12:  # регистрационный номер ккт
                    fiscal_reg_num_raw = is_kkt_num(col[row].value)
                case 13:  # оплата в такскоме до
                    fiscal_taxcom_date_raw = is_date(col[row].value)
                case 14:  # номер фн
                    fiscal_fn_num_raw = col[row].value
                case 15:  # дата окончания фн
                    fiscal_fn_end_date_raw = is_date(col[row].value)
                case 16:  # срок фн
                    fiscal_fn_period_raw = find_fn_period(col[row].value)
                case 17:  # тип компьютера на кассе
                    arm_comp_raw = col[row].value
                case 18:   # наличие ЕГАИС
                    egais_avaliable_raw = find_yes(col[row].value)
                case 19:  # срок ГОСТ ключа
                    egais_gost_key_date_raw = is_date(col[row].value)
                case 20:  # срок RSA ключа
                    egais_rsa_key_date_raw = is_date(col[row].value)
                case 21:  # fsrar id
                    egais_fsrar_id_raw = find_fsrar_id(col[row].value)
                case 22:  # ОС на кассе
                    arm_os_raw = col[row].value
                case 23:  # логический номер терминала
                    arm_pos_num_raw = find_logic_num(col[row].value)
                case 25:  # версия кассира
                    arm_strih_ver_raw = find_shtrih_ver(col[row].value)
                case 26:  # сигареты
                    cigarettes_raw = find_yes(col[row].value)
                case 27:  # считыватель пропусков
                    permit_raw = find_yes(col[row].value)

        try:
            shop_main_info = MainData(
                shop_num=shop_num_raw,
                shop_post_index=shop_post_index_raw,
                shop_address=shop_address_raw,
                shop_kpp=shop_kpp_raw,
                shop_entity=shop_entity_raw,
                shop_cigarettes=cigarettes_raw,
                shop_status=shop_status_raw
            )
        except ValidationError as e:
            print(e.json())

        try:
            fiscal_info_dump = FiscalData(
                fiscal_model=fiscal_model_raw,
                fiscal_fabric_num=fiscal_fabric_raw,
                fiscal_reg_num=fiscal_reg_num_raw,
                fascal_taxcom_name=fiscal_taxcom_name_raw,
                fiscal_taxcom_end_date=fiscal_taxcom_date_raw,
                fiscal_fn_num=fiscal_fn_num_raw,
                fiscal_fn_period=fiscal_fn_period_raw,
                fiscal_fn_end_day=fiscal_fn_end_date_raw
            )
        except ValidationError as e:
            print(e.json())

        try:
            arm_info_dump = DevicesData(
                arm_comp=arm_comp_raw,
                arm_os=arm_os_raw,
                arm_shtrih_ver=arm_strih_ver_raw,
                arm_pos_num=arm_pos_num_raw,
                arm_permit=permit_raw
            )
        except ValidationError as e:
            print(e.json())

        try:
            egais_dump = EgaisData(
                egais_avaliable=egais_avaliable_raw,
                egais_gost_key_end_date=egais_gost_key_date_raw,
                egais_rsa_key_date_raw=egais_rsa_key_date_raw,
                egais_fsrar_id=egais_fsrar_id_raw
            )
        except ValidationError as e:
            print(e.json())

        try:
            shop_info = ShopData(
                main_info=shop_main_info,
                fiscal_info=fiscal_info_dump,
                devices_info=arm_info_dump,
                egais_info=egais_dump
            )
        except ValidationError as e:
            print(e.json())

        all_shops.append(shop_info)
    return AllShopsInfo(shops=all_shops)


def monitoring_parser(file_link: str) -> AllShopsInfo:
    monitoring_book = openpyxl.load_workbook(file_link)
    worksheet = monitoring_book.active
    all_shops = []
    for row in range(18, worksheet.max_row):
        for col in worksheet.iter_cols(3, 16):
            match col[row].column:
                case 3:
                    shop_num_raw = find_shop_num(col[row].value)
                case 5:
                    shop_address_raw = col[row].value
                case 6:
                    fiscal_taxcom_name_raw = col[row].value
                case 7:
                    fiscal_reg_num_raw = is_kkt_num(col[row].value)
                case 8:
                    fiscal_fabric_num_raw = is_kkt_num(col[row].value)
                case 9:
                    fiscal_model_raw = col[row].value
                case 11:
                    fiscal_taxcom_end_date_raw = is_date(col[row].value)

        try:
            shop_main_info = MainData(
                shop_num=shop_num_raw,
                shop_address=shop_address_raw
            )
        except ValidationError as e:
            print(e.json())

        try:
            fiscal_info_dump = FiscalData(
                fiscal_model=fiscal_model_raw,
                fiscal_fabric_num=fiscal_fabric_num_raw,
                fiscal_reg_num=fiscal_reg_num_raw,
                fiscal_taxcom_name=fiscal_taxcom_name_raw,
                fiscal_taxcom_end_date=fiscal_taxcom_end_date_raw
            )
        except ValidationError as e:
            print(e.json())

        try:
            shop_info = ShopData(
                main_info=shop_main_info,
                fiscal_info=fiscal_info_dump
            )
        except ValidationError as e:
            print(e.json())
        all_shops.append(shop_info)
    return AllShopsInfo(shops=all_shops)


if __name__ == "__main__":
    # print(contacts_parser(contacts_file).json(ensure_ascii=False))
    print(alltimetable_parser(alltime_file).json(ensure_ascii=False))
    # print(monitoring_parser(monitoring_file).json(ensure_ascii=False))
